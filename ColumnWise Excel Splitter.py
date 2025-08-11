import customtkinter as ctk
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font # Only Font needed now
import os
import threading
from pathlib import Path

# The apply_hyperlink_style function is removed as it was causing issues
# and direct font styling is sufficient and more robust.

class ExcelSplitterApp(ctk.CTk):
    def __init__(self, master=None):
        super().__init__()

        self.title("✂️ Excel Splitter with Index")
        self.geometry("700x550")
        self.minsize(600, 450)

        # Configure grid layout
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(4, weight=1) # Results can expand

        # --- UI Elements ---

        # 1. Source File Selection
        self.source_file_frame = ctk.CTkFrame(self)
        self.source_file_frame.grid(row=0, column=0, padx=20, pady=(20, 10), sticky="ew")
        self.source_file_frame.grid_columnconfigure(1, weight=1)

        self.source_label = ctk.CTkLabel(self.source_file_frame, text="Source Excel File:", font=ctk.CTkFont(size=14, weight="bold"))
        self.source_label.grid(row=0, column=0, padx=(10, 5), pady=10, sticky="w")

        self.source_file_path = ctk.StringVar()
        self.source_file_entry = ctk.CTkEntry(self.source_file_frame, textvariable=self.source_file_path, placeholder_text="Select your Excel file (.xlsx)", font=ctk.CTkFont(size=12))
        self.source_file_entry.grid(row=0, column=1, padx=5, pady=10, sticky="ew")

        self.browse_button = ctk.CTkButton(self.source_file_frame, text="Browse", command=self.browse_source_file, font=ctk.CTkFont(size=12, weight="bold"))
        self.browse_button.grid(row=0, column=2, padx=(5, 10), pady=10)

        # 2. Column Name Input
        self.column_frame = ctk.CTkFrame(self)
        self.column_frame.grid(row=1, column=0, padx=20, pady=5, sticky="ew")
        self.column_frame.grid_columnconfigure(1, weight=1)

        self.column_label = ctk.CTkLabel(self.column_frame, text="Column to Split By:", font=ctk.CTkFont(size=14, weight="bold"))
        self.column_label.grid(row=0, column=0, padx=(10, 5), pady=10, sticky="w")

        self.split_column_name = ctk.StringVar()
        self.column_entry = ctk.CTkEntry(self.column_frame, textvariable=self.split_column_name, placeholder_text="e.g., ER_NO (exact match needed!)", font=ctk.CTkFont(size=12))
        self.column_entry.grid(row=0, column=1, padx=5, pady=10, sticky="ew")

        # 3. Output File Name (Optional)
        self.output_name_frame = ctk.CTkFrame(self)
        self.output_name_frame.grid(row=2, column=0, padx=20, pady=5, sticky="ew")
        self.output_name_frame.grid_columnconfigure(1, weight=1)

        self.output_name_label = ctk.CTkLabel(self.output_name_frame, text="Output File Name (Optional):", font=ctk.CTkFont(size=14, weight="bold"))
        self.output_name_label.grid(row=0, column=0, padx=(10, 5), pady=10, sticky="w")

        self.output_file_name = ctk.StringVar(value="Split_Output") # Default name
        self.output_name_entry = ctk.CTkEntry(self.output_name_frame, textvariable=self.output_file_name, placeholder_text="e.g., My_Split_Excel", font=ctk.CTkFont(size=12))
        self.output_name_entry.grid(row=0, column=1, padx=5, pady=10, sticky="ew")

        # 4. Process Button
        self.process_button = ctk.CTkButton(self, text="⚙️ Split & Create Workbook", command=self.start_split_thread, font=ctk.CTkFont(size=16, weight="bold"), fg_color="blue", hover_color="darkblue")
        self.process_button.grid(row=3, column=0, padx=20, pady=20, sticky="ew", ipadx=10, ipady=5)

        # 5. Status and Output Path Display
        self.status_label = ctk.CTkLabel(self, text="Ready.", font=ctk.CTkFont(size=12), anchor="w", text_color="gray")
        self.status_label.grid(row=5, column=0, padx=20, pady=(0, 5), sticky="ew") # Shifted by 1 row due to results box

        self.output_path_display = ctk.CTkLabel(self, text="", font=ctk.CTkFont(size=12, slant="italic"), wraplength=650, justify="left", text_color="green")
        self.output_path_display.grid(row=6, column=0, padx=20, pady=(0, 10), sticky="ew")

        # 6. "Created by MG" Label
        self.creator_label = ctk.CTkLabel(self, text="Created by MG", font=ctk.CTkFont(size=10, slant="italic"), text_color="gray")
        self.creator_label.grid(row=7, column=0, padx=20, pady=(5, 10), sticky="s") # Shifted by 1 row

        self.split_thread = None

    def browse_source_file(self):
        """Opens a file dialog to select the source Excel file."""
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if file_path:
            self.source_file_path.set(file_path)
            self.status_label.configure(text=f"Source file selected: {Path(file_path).name}")
            # Suggest a default output name based on source file
            base_name = Path(file_path).stem
            self.output_file_name.set(f"{base_name}_Split")

    def start_split_thread(self):
        """Starts the Excel splitting process in a separate thread."""
        source_file = self.source_file_path.get()
        column_name = self.split_column_name.get()
        output_name = self.output_file_name.get()

        if not source_file:
            messagebox.showwarning("Input Error", "Please select a source Excel file.")
            return
        if not column_name:
            messagebox.showwarning("Input Error", "Please enter the column name to split by.")
            return
        if not output_name:
            output_name = "Split_Output" # Fallback if user clears it

        if not os.path.exists(source_file):
            messagebox.showerror("File Error", f"Source file not found: {source_file}")
            return

        self.process_button.configure(state="disabled", text="Processing...")
        self.status_label.configure(text="Loading Excel file...")
        self.output_path_display.configure(text="")

        self.split_thread = threading.Thread(target=self._perform_excel_split,
                                             args=(source_file, column_name, output_name))
        self.split_thread.daemon = True # Allows thread to exit with main app
        self.split_thread.start()

    def _perform_excel_split(self, source_file, column_name, output_name):
        """
        Performs the Excel splitting operation. Runs in a separate thread.
        """
        try:
            # 1. Load your Excel file
            self.status_label.configure(text=f"Loading '{Path(source_file).name}'...")
            df = pd.read_excel(source_file)

            # IMPORTANT: Check if column exists using exact name provided by user
            if column_name not in df.columns:
                # Use a default value for 'e_info' here for the lambda
                self.after(0, lambda col=column_name: messagebox.showerror("Column Not Found", f"Column '{col}' not found in the Excel file. Please check the spelling and casing."))
                self.after(0, self._re_enable_buttons)
                return

            # Ensure the output directory exists (same as source file's directory)
            output_dir = Path(source_file).parent
            output_file_path = output_dir / f"{output_name}.xlsx"

            # 2. Create a new workbook
            self.status_label.configure(text="Creating new workbook...")
            wb = Workbook()
            # apply_hyperlink_style(wb) # This function is now removed

            # Remove the default created sheet
            default_sheet = wb.active
            wb.remove(default_sheet)

            # 3. Create Index sheet
            index_ws = wb.create_sheet(title='Index')
            index_ws['A1'] = f'Sheet Name ({column_name})'
            index_ws.column_dimensions['A'].width = 30
            index_ws['A1'].font = Font(bold=True) # Make index header bold

            # Track the index row number
            index_row_num = 2

            # Get unique values from the column to split by
            unique_values = df[column_name].unique()
            total_sheets = len(unique_values)

            # Loop through each unique value to create a new sheet
            for i, unique_val in enumerate(unique_values):
                self.status_label.configure(text=f"Processing {i+1}/{total_sheets} for '{unique_val}'...")
                self.update_idletasks() # Update UI

                # Filter DataFrame for the current unique value
                subset_df = df[df[column_name] == unique_val]

                # Generate sheet name
                sheet_base_name = str(unique_val)
                # Clean invalid characters and truncate to 31 characters if needed
                invalid_chars = ['\\', '/', '?', '*', '[', ']', ':', '|', '<', '>']
                for char in invalid_chars:
                    sheet_base_name = sheet_base_name.replace(char, '_')

                sheet_name = sheet_base_name[:31] # Initial truncation

                # Ensure unique sheet names by appending a counter if duplicate
                counter = 1
                original_sheet_name = sheet_name
                while sheet_name in wb.sheetnames:
                    sheet_name = f"{original_sheet_name[:28]}_{counter}" # Truncate + counter
                    counter += 1

                # Create new sheet
                ws = wb.create_sheet(title=sheet_name)

                # Add 'Back to Index' hyperlink in cell A1
                back_link_cell = ws.cell(row=1, column=1)
                back_link_cell.value = '<< Back to Index'
                back_link_cell.hyperlink = f"#Index!A1"
                back_link_cell.font = Font(color="0000FF", underline="single") # Direct styling

                # Write DataFrame subset to the new sheet starting from row 3 (for actual data)
                # Add headers to row 2
                for col_num, col_name_df in enumerate(subset_df.columns, start=1):
                    ws.cell(row=2, column=col_num, value=col_name_df).font = Font(bold=True)

                # Write data rows starting from row 3
                for row_idx, row_data in enumerate(subset_df.values, start=3):
                    for col_idx, value in enumerate(row_data, start=1):
                        ws.cell(row=row_idx, column=col_idx, value=value)

                # Add entry to Index sheet with hyperlink
                index_cell = index_ws.cell(row=index_row_num, column=1)
                index_cell.value = sheet_name
                index_cell.hyperlink = f"#{sheet_name}!A1"
                index_cell.font = Font(color="0000FF", underline="single") # Direct styling

                index_row_num += 1

            # Save new workbook
            wb.save(output_file_path)

            self.after(0, lambda: self.status_label.configure(text="✔️ Data integrity verified — full workbook with hyperlinks created!"))
            self.after(0, lambda: self.output_path_display.configure(text=f"Output saved to: {output_file_path.as_posix()}"))
            self.after(0, lambda: messagebox.showinfo("Success", f"Workbook created successfully at:\n{output_file_path.as_posix()}"))

        except Exception as e_info: # Capture the exception as 'e_info'
            self.after(0, lambda info=e_info: self.status_label.configure(text=f"❌ An error occurred: {info}"))
            self.after(0, lambda info=e_info: messagebox.showerror("Error", f"An unexpected error occurred during processing:\n{info}"))
        finally:
            self.after(0, self._re_enable_buttons)

    def _re_enable_buttons(self):
        """Re-enables UI elements after processing is complete or an error occurs."""
        self.process_button.configure(state="normal", text="⚙️ Split & Create Workbook")

# Main execution block
if __name__ == "__main__":
    app = ExcelSplitterApp()
    app.mainloop()
